import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import load_workbook
import csv
import requests
import random
import string

class CreateBranchApp(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.configure(bg="#d9e6fc")
        self.pack(fill="both", expand=True)

        lbl_title = tk.Label(self, text="Create Branch", fg='black', bg="#d9e6fc",
                             font=('cambria', 20, 'bold'), padx=10, pady=10)
        lbl_title.pack(side=tk.TOP, fill=tk.X)

        self.form_frame = tk.Frame(self, padx=20, pady=20, bg="#d9e6fc")
        self.form_frame.pack()

        self.create_widgets()

    def create_widgets(self):
        font_label = ('calibri', 15)
        font_entry = ('calibri', 15)

        # Branch Name
        tk.Label(self.form_frame, text="Branch Name:", bg="#d9e6fc", font=font_label).grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)
        self.ent_name = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_name.grid(row=0, column=1, padx=10, pady=10)

        # Branch City
        tk.Label(self.form_frame, text="City:", bg="#d9e6fc", font=font_label).grid(row=0, column=2, padx=10, pady=10, sticky=tk.W)
        self.ent_city = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_city.grid(row=0, column=3, padx=10, pady=10)

        # State Code
        tk.Label(self.form_frame, text="State Code:", bg="#d9e6fc", font=font_label).grid(row=1, column=0, padx=10, pady=10, sticky=tk.W)
        self.ent_statec = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_statec.grid(row=1, column=1, padx=10, pady=10)
        self.ent_statec.bind("<FocusOut>", self.fetch_state_name)
        self.ent_statec.bind("<Return>", self.fetch_state_name)

        # State Name (readonly)
        tk.Label(self.form_frame, text="State Name:", bg="#d9e6fc", font=font_label).grid(row=1, column=2, padx=10, pady=10, sticky=tk.W)
        self.ent_staten = tk.Entry(self.form_frame, font=font_entry, width=25, state="readonly")
        self.ent_staten.grid(row=1, column=3, padx=10, pady=10)

        # Buttons
        btn_frame = tk.Frame(self, bg="#d9e6fc")
        btn_frame.pack(pady=20)

        tk.Button(btn_frame, text="Create Branch", width=20, bg="#1a5dd1", fg="white",
                  font=("Mulish", 11, "bold"), command=self.create_branch).grid(row=0, column=0, padx=10)
        tk.Button(btn_frame, text="Upload Excel / CSV", width=20, bg="#194a7a", fg="white",
                  font=("Mulish", 11, "bold"), command=self.upload_file).grid(row=0, column=1, padx=10)
        tk.Button(btn_frame, text="Clear", width=20, bg="#6C757D", fg="white",
                  font=("Mulish", 11, "bold"), command=self.clear_fields).grid(row=0, column=2, padx=10)

    def clear_fields(self):
        for widget in [self.ent_name, self.ent_city, self.ent_statec, self.ent_staten]:
            widget.config(state="normal")
            widget.delete(0, tk.END)
        self.ent_staten.config(state="readonly")

    def fetch_state_name(self, event=None):
        state_code = self.ent_statec.get().strip()
        if not state_code:
            return
        try:
            response = requests.get("http://127.0.0.1:5000/get_state/search", params={"query": state_code})
            result = response.json()

            self.ent_staten.config(state="normal")
            self.ent_staten.delete(0, tk.END)

            if response.status_code == 200 and result.get("success"):
                first = result["states"][0]
                self.ent_staten.insert(0, first["state_Name"])
            else:
                messagebox.showwarning("Not Found", "State code not found.")
            self.ent_staten.config(state="readonly")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to fetch state name:\n{e}")

    def generate_branch_code(self, name, city):
        name_part = (name[:3].upper() if len(name) >= 3 else name.upper().ljust(3, 'X'))
        city_part = (city[:2].upper() if len(city) >= 2 else city.upper().ljust(2, 'Y'))
        rand_part = ''.join(random.choices(string.digits, k=2))
        return (name_part + city_part + rand_part)[:7]

    def create_branch(self):
        name = self.ent_name.get().strip()
        city = self.ent_city.get().strip()
        statec = self.ent_statec.get().strip()
        staten = self.ent_staten.get().strip()

        if not all([name, city, statec, staten]):
            messagebox.showwarning("Validation Error", "All fields are required.")
            return

        branch_code = self.generate_branch_code(name, city)

        payload = {
            "branch_code": branch_code,
            "branch_name": name,
            "branch_city": city,
            "branch_statec": statec,
            "branch_staten": staten
        }

        try:
            response = requests.post("http://127.0.0.1:5000/create_branch", json=payload)
            res = response.json()
            if res.get("success"):
                messagebox.showinfo("Success", f"Branch created successfully with Code: {branch_code}")
                self.clear_fields()
            else:
                messagebox.showerror("Error", res.get("message", "Unknown error"))
        except Exception as e:
            messagebox.showerror("Server Error", f"Failed to create branch: {e}")

    def upload_file(self):
        filetypes = [("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
        filename = filedialog.askopenfilename(title="Select Excel/CSV file", filetypes=filetypes)
        if not filename:
            return

        try:
            if filename.endswith(".xlsx"):
                wb = load_workbook(filename)
                ws = wb.active
                data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
            else:
                with open(filename, newline='') as csvfile:
                    reader = csv.reader(csvfile)
                    next(reader)  # skip header
                    data = [row for row in reader]

            count = 0
            for row in data:
                if len(row) < 3:
                    continue
                branch_name, branch_city, branch_statec = row[:3]
                branch_staten = self.get_state_name(branch_statec)
                if not branch_staten:
                    continue

                branch_code = self.generate_branch_code(branch_name, branch_city)

                payload = {
                    "branch_code": branch_code,
                    "branch_name": branch_name,
                    "branch_city": branch_city,
                    "branch_statec": branch_statec,
                    "branch_staten": branch_staten
                }

                response = requests.post("http://127.0.0.1:5000/create_branch", json=payload)
                if response.status_code == 200 and response.json().get("success"):
                    count += 1

            messagebox.showinfo("Upload Result", f"{count} branches uploaded successfully.")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to upload file:\n{e}")

    def get_state_name(self, code):
        try:
            response = requests.get("http://127.0.0.1:5000/get_state/search", params={"query": code})
            result = response.json()
            if response.status_code == 200 and result.get("success"):
                return result["states"][0]["state_Name"]
        except:
            pass
        return None

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Branch Creation")
    root.geometry("1000x500")
    app = CreateBranchApp(root)
    root.mainloop()
