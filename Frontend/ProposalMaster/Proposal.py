import tkinter as tk
from tkinter import *
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook, load_workbook
from datetime import datetime
import csv
import random
import mysql.connector

class ProposalApp(tk.Frame):
    def __init__(self, parent, home_page=None):
        super().__init__(parent)
        self.parent = parent
        self.home_page = home_page

        # ---------- Title ----------
        lbl_title = tk.Label(self, text="Proposal Master", fg='black',
                             font=('cambria', 20, 'bold'), padx=10, pady=10)
        lbl_title.pack(side=tk.TOP, fill=tk.X)

        # ---------- Form Frame ----------
        self.form_frame = tk.Frame(self, padx=20, pady=20)
        self.form_frame.pack()

        # Create widgets
        self.create_widgets()

    def create_widgets(self):
        font_label = ('calibri', 14)
        font_entry = ('calibri', 14)

        # Row 0 - Customer ID and Name
        tk.Label(self.form_frame, text="Customer ID:", font=font_label).grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)
        self.ent_cust_id = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_cust_id.grid(row=0, column=1, padx=10, pady=10)
        self.ent_cust_id.bind("<FocusOut>", self.fetch_customer_name)
        
        # Bind Enter key to trigger name fetch
        self.ent_cust_id.bind("<Return>", self.fetch_customer_name)
    
        tk.Label(self.form_frame, text="Customer Name:", font=font_label).grid(row=0, column=2, padx=10, pady=10, sticky=tk.W)
        self.ent_cust_name = tk.Entry(self.form_frame, font=font_entry, width=25, state="readonly")
        self.ent_cust_name.grid(row=0, column=3, padx=10, pady=10)

        # Row 1 - Proposal Date and Premium
        tk.Label(self.form_frame, text="Proposal Date (DD/MM/YYYY):", font=font_label).grid(row=1, column=0, padx=10, pady=10, sticky=tk.W)
        self.ent_proposal_date = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_proposal_date.grid(row=1, column=1, padx=10, pady=10)

        tk.Label(self.form_frame, text="Premium:", font=font_label).grid(row=1, column=2, padx=10, pady=10, sticky=tk.W)
        self.ent_premium = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_premium.grid(row=1, column=3, padx=10, pady=10)

        # Row 2 - Premium Term
        tk.Label(self.form_frame, text="Premium Term:", font=font_label).grid(row=2, column=0, padx=10, pady=10, sticky=tk.W)
        self.ent_premium_term = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_premium_term.grid(row=2, column=1, padx=10, pady=10)

        # Row 3 - Employee ID and Name
        tk.Label(self.form_frame, text="Employee ID:", font=font_label).grid(row=3, column=0, padx=10, pady=10, sticky=tk.W)
        self.ent_emp_id = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_emp_id.grid(row=3, column=1, padx=10, pady=10)
        self.ent_emp_id.bind("<FocusOut>", self.fetch_employee_name)

        tk.Label(self.form_frame, text="Employee Name:", font=font_label).grid(row=3, column=2, padx=10, pady=10, sticky=tk.W)
        self.ent_emp_name = tk.Entry(self.form_frame, font=font_entry, width=25, state="readonly")
        self.ent_emp_name.grid(row=3, column=3, padx=10, pady=10)

        # Row 4 - CP Code and Name
        tk.Label(self.form_frame, text="Channel Partner Code:", font=font_label).grid(row=4, column=0, padx=10, pady=10, sticky=tk.W)
        self.ent_cp_code = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_cp_code.grid(row=4, column=1, padx=10, pady=10)
        self.ent_cp_code.bind("<FocusOut>", self.fetch_cp_name)
        
        # Bind Enter key to trigger name fetch
        self.ent_cp_code.bind("<Return>", self.fetch_cp_name)

        tk.Label(self.form_frame, text="Partner Name:", font=font_label).grid(row=4, column=2, padx=10, pady=10, sticky=tk.W)
        self.ent_cp_name = tk.Entry(self.form_frame, font=font_entry, width=25, state="readonly")
        self.ent_cp_name.grid(row=4, column=3, padx=10, pady=10)

        # Row 5 - Business Type and Premium Mode
        tk.Label(self.form_frame, text="Business Type:", font=font_label).grid(row=5, column=0, padx=10, pady=10, sticky=tk.W)
        self.cmb_business_type = ttk.Combobox(self.form_frame, font=font_entry, width=25, values=["NB", "FYIP", "Renewal"])
        self.cmb_business_type.grid(row=5, column=1, padx=10, pady=10)

        tk.Label(self.form_frame, text="Premium Mode:", font=font_label).grid(row=5, column=2, padx=10, pady=10, sticky=tk.W)
        self.cmb_premium_mode = ttk.Combobox(self.form_frame, font=font_entry, width=25, values=["Yearly", "Half Yearly", "Quarterly", "Monthly"])
        self.cmb_premium_mode.grid(row=5, column=3, padx=10, pady=10)

        # ---------- Buttons ----------
        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=15)

        tk.Button(btn_frame, text="Create Proposal", width=20, height=1, fg='#dddddd', bg='#ff3300',
                  font=('calibri', 12, 'bold'), command=self.create_proposal).grid(row=0, column=0, padx=10)

        tk.Button(btn_frame, text="Upload Excel/CSV", width=20, height=1, fg='#dddddd', bg='#ff3300',
                  font=('calibri', 12, 'bold'), command=self.upload_file).grid(row=0, column=1, padx=10)

        tk.Button(btn_frame, text="Clear", width=20, height=1, fg='#dddddd', bg='#ff3300',
                  font=('calibri', 12, 'bold'), command=self.clear_fields).grid(row=0, column=2, padx=10)

    def clear_fields(self):
        for widget in [self.ent_cust_id, self.ent_cust_name, self.ent_proposal_date, self.ent_premium,
                       self.ent_premium_term, self.ent_emp_id, self.ent_emp_name,
                       self.ent_cp_code, self.ent_cp_name]:
            widget.config(state="normal")
            widget.delete(0, tk.END)
            if "name" in str(widget).lower():
                widget.config(state="readonly")

        self.cmb_business_type.set("")
        self.cmb_premium_mode.set("")

    def fetch_customer_name(self, event=None):
        cust_id = self.ent_cust_id.get().strip()
        
        try:
            connection = mysql.connector.connect(
                host="localhost",
                user="root",
                password="@Raja12",
                database="lidata"
            )
            cursor = connection.cursor()
            cursor.execute("SELECT cust_name FROM customer WHERE cust_id = %s", (cust_id,))
            result = cursor.fetchone()

            self.ent_cust_name.config(state="normal")
            self.ent_cust_name.delete(0, tk.END)

            if result:
                self.ent_cust_name.insert(0, result[0])
            else:
                self.ent_cust_name.insert(0, "")  # Not found
                self.ent_cust_name.config(state="readonly")
                cursor.close()
                connection.close()
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", f"Failed to fetch customer name:\n{err}")

    def fetch_employee_name(self, event=None):
        employee_data = {"E001": "Alice Brown", "E002": "Bob White"}
        emp_id = self.ent_emp_id.get().strip()
        self.ent_emp_name.config(state="normal")
        self.ent_emp_name.delete(0, tk.END)
        self.ent_emp_name.insert(0, employee_data.get(emp_id, ""))
        self.ent_emp_name.config(state="readonly")

    def fetch_cp_name(self, event=None):
        cp_code = self.ent_cp_code.get().strip()

        try:
            connection = mysql.connector.connect(
                host="localhost",
                user="root",
                password="@Raja12",
                database="lidata"
            )
            cursor = connection.cursor()
            cursor.execute("SELECT CP_Name FROM cp WHERE CP_Code = %s", (cp_code,))
            result = cursor.fetchone()

            self.ent_cp_name.config(state="normal")
            self.ent_cp_name.delete(0, tk.END)

            if result:
                self.ent_cp_name.insert(0, result[0])
            else:
                self.ent_cp_name.insert(0, "")  # Not found
                self.ent_cp_name.config(state="readonly")
                cursor.close()
                connection.close()
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", f"Failed to fetch CP name:\n{err}")

    def generate_proposal_number(self):
        prefix = "5000"
        while True:
            random_digits = ''.join([str(random.randint(0, 9)) for _ in range(8)])
            proposal_number = prefix + random_digits
            if not self.proposal_number_exists(proposal_number):
                return proposal_number

    def proposal_number_exists(self, proposal_number):
        try:
            connection = mysql.connector.connect(
                host="localhost",
                user="root",
                password="@Raja12",
                database="lidata"
            )
            cursor = connection.cursor()
            query = "SELECT 1 FROM proposal WHERE proposal_number = %s"
            cursor.execute(query, (proposal_number,))
            exists = cursor.fetchone() is not None
            cursor.close()
            connection.close()
            return exists
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", f"Failed to check proposal number:\n{err}")
            return True  # assume duplicate to be safe
    
    def create_proposal(self):
        if not all([self.ent_cust_id.get(), self.ent_proposal_date.get(), self.ent_premium.get(),
                    self.ent_premium_term.get(), self.ent_emp_id.get(), self.ent_cp_code.get(),
                    self.cmb_business_type.get(), self.cmb_premium_mode.get()]):
            messagebox.showerror("Validation Error", "All fields must be filled out.")
            return

        try:
            datetime.strptime(self.ent_proposal_date.get(), "%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Validation Error", "Invalid date format. Use DD/MM/YYYY.")
            return

        proposal_number = self.generate_proposal_number()

        file_path = "proposals.xlsx"
        try:
            wb = load_workbook(file_path)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(["Proposal Number", "Customer ID", "Customer Name", "Proposal Date", "Premium", "Premium Term",
                       "Employee ID", "Employee Name", "Channel Partner Code", "Channel Partner Name",
                       "Business Type", "Premium Mode"])

        proposal_data = [
            proposal_number,
            self.ent_cust_id.get(),
            self.ent_cust_name.get(),
            self.ent_proposal_date.get(),
            self.ent_premium.get(),
            self.ent_premium_term.get(),
            self.ent_emp_id.get(),
            self.ent_emp_name.get(),
            self.ent_cp_code.get(),
            self.ent_cp_name.get(),
            self.cmb_business_type.get(),
            self.cmb_premium_mode.get()
        ]

        ws.append(proposal_data)
        wb.save(file_path)

        self.insert_into_database(proposal_data)

        messagebox.showinfo("Success", f"Proposal #{proposal_number} saved successfully!")
        self.clear_fields()

    def insert_into_database(self, data):
        try:
            connection = mysql.connector.connect(
                host="localhost",
                user="root",
                password="@Raja12",
                database="lidata"
            )
            cursor = connection.cursor()

            query = """
                INSERT INTO proposal (
                    proposal_number, customer_id, customer_name, proposal_date, premium, premium_term,
                    employee_id, employee_name, channel_partner_code, channel_partner_name,
                    business_type, premium_mode
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """

            cursor.execute(query, data)
            connection.commit()
            cursor.close()
            connection.close()
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", f"Failed to insert into database:\n{err}")

    def upload_file(self):
        filetypes = [("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
        filename = filedialog.askopenfilename(title="Select Excel/CSV file", filetypes=filetypes)
        if not filename:
            return

        try:
            if filename.endswith(".xlsx"):
                wb = load_workbook(filename)
                ws = wb.active
                data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
            else:
                with open(filename, newline='') as csvfile:
                    reader = csv.reader(csvfile)
                    next(reader)
                    data = [row for row in reader]

            messagebox.showinfo("File Loaded", f"Loaded {len(data)} rows from file: {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file:\n{e}")


if __name__ == "__main__":
    root = tk.Tk()
    app = ProposalApp(root)
    root.mainloop()