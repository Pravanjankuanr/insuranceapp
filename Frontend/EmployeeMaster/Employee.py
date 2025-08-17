import tkinter as tk
from tkinter import *
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook, load_workbook
from datetime import datetime
import csv
import random
import mysql.connector

class EmployeeApp(tk.Frame):
    def __init__(self, parent, home_page=None):
        super().__init__(parent)
        self.parent = parent
        self.home_page = home_page        

        # ---------- Title ----------
        lbl_title = tk.Label(self, text="Employee Master", fg='black',
                             font=('cambria', 20, 'bold'), padx=10, pady=10)
        lbl_title.pack(side=tk.TOP, fill=tk.X)

        # ---------- Form Frame ----------
        self.form_frame = tk.Frame(self, padx=20, pady=20)
        self.form_frame.pack()

        # Create widgets
        self.create_widgets()

    def create_widgets(self):
        font_label = ('calibri', 15)
        font_entry = ('calibri', 15)

        # Row 0 - Customer Name
        tk.Label(self.form_frame, text="Employee Name:", font=font_label).grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)
        self.ent_emp_name = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_emp_name.grid(row=0, column=1, padx=10, pady=10)
        
        # Row 3 - Branch Code and Name
        tk.Label(self.form_frame, text="Branch Code:", font=font_label).grid(row=3, column=0, padx=10, pady=10, sticky=tk.W)
        self.ent_branchc = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_branchc.grid(row=3, column=1, padx=10, pady=10)
        self.ent_branchc.bind("<FocusOut>", self.fetch_branchn)
        
        # Bind Enter key to trigger name fetch
        self.ent_branchc.bind("<Return>", self.fetch_branchn)

        tk.Label(self.form_frame, text="Branch Name:", font=font_label).grid(row=3, column=2, padx=10, pady=10, sticky=tk.W)
        self.ent_branchn = tk.Entry(self.form_frame, font=font_entry, width=25, state="readonly")
        self.ent_branchn.grid(row=3, column=3, padx=10, pady=10)


        # ---------- Buttons ----------
        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=15)

        tk.Button(btn_frame, text="Create Employee", width=20, height=1, fg='#dddddd', bg='#ff3300',
                  font=('calibri', 12, 'bold'), command=self.create_employee).grid(row=0, column=0, padx=10)

        tk.Button(btn_frame, text="Upload Excel/CSV", width=20, height=1, fg='#dddddd', bg='#ff3300',
                  font=('calibri', 12, 'bold'), command=self.upload_file).grid(row=0, column=1, padx=10)

        tk.Button(btn_frame, text="Clear", width=20, height=1, fg='#dddddd', bg='#ff3300',
                  font=('calibri', 12, 'bold'), command=self.clear_fields).grid(row=0, column=2, padx=10)

    def clear_fields(self):
        for widget in [self.ent_emp_name, self.ent_branchc, self.ent_branchc]:
            widget.config(state="normal")
            widget.delete(0, tk.END)
            if "name" in str(widget).lower():
                widget.config(state="readonly")

    def fetch_branchn(self, event=None):
        branchc = self.ent_branchc.get().strip()

        try:
            connection = mysql.connector.connect(
                host="localhost",
                user="root",
                password="@Raja12",
                database="lidata"
            )
            cursor = connection.cursor()
            cursor.execute("SELECT branch_name FROM branch WHERE branch_code = %s", (branchc,))
            result = cursor.fetchone()

            self.ent_branchn.config(state="normal")
            self.ent_branchn.delete(0, tk.END)

            if result:
                self.ent_branchn.insert(0, result[0])
            else:
                self.ent_branchn.insert(0, "")  # Not found
                self.ent_branchn.config(state="readonly")
                cursor.close()
                connection.close()
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", f"Failed to fetch branch name:\n{err}")

    def generate_employee_id(self):
        prefix = "LI"
        while True:
            random_digits = ''.join([str(random.randint(0, 9)) for _ in range(4)])
            employee_id = prefix + random_digits
            if not self.employee_id_exists(employee_id):
                return employee_id

    def employee_id_exists(self, emp_id):
        try:
            connection = mysql.connector.connect(
                host="localhost",
                user="root",
                password="@Raja12",
                database="lidata"
            )
            cursor = connection.cursor()
            query = "SELECT 1 FROM employee WHERE emp_id = %s"
            cursor.execute(query, (emp_id,))
            exists = cursor.fetchone() is not None
            cursor.close()
            connection.close()
            return exists
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", f"Failed to check employee id:\n{err}")
            return True
    
    def create_employee(self):
        if not all([self.ent_emp_name.get(), self.ent_branchc.get(), self.ent_branchn.get()]):
            messagebox.showerror("Validation Error", "All fields must be filled out.")
            return

        emp_id = self.generate_employee_id()

        file_path = "employee.xlsx"
        try:
            wb = load_workbook(file_path)
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(["Employee Name", "Branch Code", "Branch Name"])

        Employee_data = [
            emp_id,
            self.ent_emp_name.get(),
            self.ent_branchc.get(),
            self.ent_branchn.get()
        ]

        ws.append(Employee_data)
        wb.save(file_path)
    
        self.insert_into_database(Employee_data)

        messagebox.showinfo("Success", f"Customer id #{emp_id} saved successfully!")
        # self.clear_fields()

    def insert_into_database(self, data):
        try:
            connection = mysql.connector.connect(
                host="localhost",
                user="root",
                password="@Raja12",
                database="lidata"
            )
            cursor = connection.cursor()

            query = """
                INSERT INTO employee (
                    emp_id, emp_name, branch_code, branch_name
                ) VALUES (%s, %s, %s, %s)
            """

            cursor.execute(query, data)
            connection.commit()
            cursor.close()
            connection.close()
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", f"Failed to insert into database:\n{err}")

    def upload_file(self):
        filetypes = [("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
        filename = filedialog.askopenfilename(title="Select Excel/CSV file", filetypes=filetypes)
        if not filename:
            return

        try:
            if filename.endswith(".xlsx"):
                wb = load_workbook(filename)
                ws = wb.active
                data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
            else:
                with open(filename, newline='') as csvfile:
                    reader = csv.reader(csvfile)
                    next(reader)
                    data = [row for row in reader]

            messagebox.showinfo("File Loaded", f"Loaded {len(data)} rows from file: {filename}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load file:\n{e}")
            
if __name__ == "__main__":
    root = tk.Tk()
    app = EmployeeApp(root)
    root.mainloop()