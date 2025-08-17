import tkinter as tk
from tkinter import *
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook, load_workbook
from datetime import datetime
import csv
import random
import mysql.connector
import requests

class CustomerApp(tk.Frame):
    def __init__(self, parent, home_page=None):
        super().__init__(parent)
        self.parent = parent
        self.home_page = home_page
        self.configure(bg="#d9e6fc")
                

        # ---------- Title ----------
        lbl_title = tk.Label(self, text="Create Customer", fg='black',bg="#d9e6fc",
                             font=('cambria', 20, 'bold'), padx=10, pady=10)
        lbl_title.pack(side=tk.TOP, fill=tk.X)

        # ---------- Form Frame ----------
        self.form_frame = tk.Frame(self, padx=20, pady=20, bg="#d9e6fc")
        self.form_frame.pack()

        # Create widgets
        self.create_widgets()

    def create_widgets(self):
        font_label = ('calibri', 15)
        font_entry = ('calibri', 15)

        # Row 0 - Customer Name
        tk.Label(self.form_frame, text="Customer Name:",bg="#d9e6fc", font=font_label).grid(row=0, column=0, padx=10, pady=10, sticky=tk.W)
        self.ent_cust_name = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_cust_name.grid(row=0, column=1, padx=10, pady=10)
        
        
        # Row 0 - Customer Number
        tk.Label(self.form_frame, text="Mobile:",bg="#d9e6fc", font=font_label).grid(row=0, column=2, padx=10, pady=10, sticky=tk.W)
        self.ent_cust_con = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_cust_con.grid(row=0, column=3, padx=10, pady=10)

        # Row 1 - Customer DOB
        tk.Label(self.form_frame, text="DOB (DD/MM/YYYY):",bg="#d9e6fc", font=font_label).grid(row=1, column=0, padx=10, pady=10, sticky=tk.W)
        self.ent_cust_dob = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_cust_dob.grid(row=1, column=1, padx=10, pady=10)

        # Row 1 - Customer PAN
        tk.Label(self.form_frame, text="PAN Number:",bg="#d9e6fc", font=font_label).grid(row=1, column=2, padx=10, pady=10, sticky=tk.W)
        self.ent_cust_pan = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_cust_pan.grid(row=1, column=3, padx=10, pady=10)

        # Row 2 - Customer Email
        tk.Label(self.form_frame, text="Email:",bg="#d9e6fc", font=font_label).grid(row=2, column=0, padx=10, pady=10, sticky=tk.W)
        self.ent_cust_email = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_cust_email.grid(row=2, column=1, padx=10, pady=10)

        # Row 2 - Customer Address
        tk.Label(self.form_frame, text="Address:",bg="#d9e6fc", font=font_label).grid(row=2, column=2, padx=10, pady=10, sticky=tk.W)
        self.ent_cust_add = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_cust_add.grid(row=2, column=3, padx=10, pady=10)

        # Row 3 - State Code and Name
        tk.Label(self.form_frame, text="State Code:",bg="#d9e6fc", font=font_label).grid(row=3, column=0, padx=10, pady=10, sticky=tk.W)
        self.ent_statec = tk.Entry(self.form_frame, font=font_entry, width=25)
        self.ent_statec.grid(row=3, column=1, padx=10, pady=10)
        self.ent_statec.bind("<FocusOut>", self.fetch_staten)
        
        # Bind Enter key to trigger name fetch
        self.ent_statec.bind("<Return>", self.fetch_staten)

        tk.Label(self.form_frame, text="State Name:",bg="#d9e6fc", font=font_label).grid(row=3, column=2, padx=10, pady=10, sticky=tk.W)
        self.ent_staten = tk.Entry(self.form_frame,bg="#d9e6fc", font=font_entry, width=25, state="readonly")
        self.ent_staten.grid(row=3, column=3, padx=10, pady=10)


        # ---------- Buttons ----------
        btn_frame = tk.Frame(self,bg="#d9e6fc")
        btn_frame.pack(pady=15)

        tk.Button(btn_frame, text="Create Customer", width=20, height=1, fg='#ffffff', bg='#1a5dd1',
                  font=('Mulish', 11, 'bold'), command=self.create_customer).grid(row=0, column=0, padx=10)

        tk.Button(btn_frame, text="Upload Excel / CSV", width=20, height=1, fg='#ffffff', bg='#194a7a',
                  font=('Mulish', 11, 'bold'), command=self.upload_file).grid(row=0, column=1, padx=10)

        tk.Button(btn_frame, text="Clear", width=20, height=1, fg='#ffffff', bg='#6C757D',
                  font=('Mulish', 11, 'bold'), command=self.clear_fields).grid(row=0, column=2, padx=10)

    def clear_fields(self):
        for widget in [self.ent_cust_name, self.ent_cust_con, self.ent_cust_dob, self.ent_cust_pan,
                        self.ent_cust_email, self.ent_cust_add, self.ent_statec, self.ent_staten]:
            widget.config(state="normal")
            widget.delete(0, tk.END)
            if "name" in str(widget).lower():
                widget.config(state="readonly")
        self.ent_staten.config(state="readonly")

    def fetch_staten(self, event=None):
        statec = self.ent_statec.get().strip()

        if not statec:
            return

        try:
            response = requests.get("http://127.0.0.1:5000/get_state/search", params={"query": statec})
            result = response.json()

            self.ent_staten.config(state="normal")
            self.ent_staten.delete(0, tk.END)

            if response.status_code == 200 and result.get("success"):
                first_result = result["states"][0]
                self.ent_staten.insert(0, first_result["state_Name"])
                self.ent_staten.config(state="readonly")

            else:
                self.ent_staten.insert(0, "")
                self.ent_staten.config(state="readonly")
                messagebox.showwarning("Not Found", "State code not found.")

        except requests.exceptions.RequestException as e:
            messagebox.showerror("Connection Error", f"Failed to connect to server:\n{e}")

    def generate_customer_id(self):
        prefix = "1011"
        while True:
            random_digits = ''.join([str(random.randint(0, 9)) for _ in range(5)])
            customer_id = prefix + random_digits
            if not self.customer_id_exists(customer_id):
                return customer_id

    def customer_id_exists(self, cust_id):
        try:
            connection = mysql.connector.connect(
                host="localhost",
                user="root",
                password="@Raja12",
                database="lidata"
            )
            cursor = connection.cursor()
            query = "SELECT 1 FROM customer WHERE cust_id = %s"
            cursor.execute(query, (cust_id,))
            exists = cursor.fetchone() is not None
            cursor.close()
            connection.close()
            return exists
        except mysql.connector.Error as err:
            messagebox.showerror("Database Error", f"Failed to check customer id:\n{err}")
            return True
    
    def create_customer(self):
        # Validate all fields
        if not all([self.ent_cust_name.get(), self.ent_cust_con.get(), self.ent_cust_dob.get(),
                    self.ent_cust_pan.get(), self.ent_cust_email.get(), self.ent_cust_add.get(),
                    self.ent_statec.get(), self.ent_staten.get()]):
            messagebox.showerror("Validation Error", "All fields must be filled out.")
            return

        # Validate DOB format
        try:
            datetime.strptime(self.ent_cust_dob.get(), "%d/%m/%Y")
        except ValueError:
            messagebox.showerror("Validation Error", "Invalid date format. Use DD/MM/YYYY.")
            return

        cust_id = self.generate_customer_id()

        payload = {
            "cust_id": cust_id,
            "cust_name": self.ent_cust_name.get(),
            "mobile": self.ent_cust_con.get(),
            "dob": self.ent_cust_dob.get(),
            "pan": self.ent_cust_pan.get(),
            "email": self.ent_cust_email.get(),
            "address": self.ent_cust_add.get(),
            "state_code": self.ent_statec.get(),
            "state_name": self.ent_staten.get()
        }

        try:
            response = requests.post("http://127.0.0.1:5000/create_customer", json=payload)
            result = response.json()

            if response.status_code == 200:
                messagebox.showinfo("Success", result.get("message", "Customer added!"))
                # self.clear_fields()
            else:
                messagebox.showerror("Error", result.get("error", "Failed to create customer."))
        except requests.exceptions.RequestException as e:
            messagebox.showerror("Connection Error", f"Could not reach server:\n{e}")

    def get_state_name_by_code(self, code):
        try:
            response = requests.get(f"http://127.0.0.1:5000/get_state?code={code}")
            result = response.json()
            if result.get("success"):
                return result.get("state_name")
            else:
                return None
        except Exception as e:
            messagebox.showerror("Error", f"Failed to fetch state name for {code}:\n{e}")
            return None

    def insert_customer_api(self, data):
        try:
            response = requests.post("http://127.0.0.1:5000/create_customer", json=data)
            result = response.json()
            if response.status_code == 200:
                return True
            else:
                return False
        except Exception as e:
            messagebox.showerror("API Error", f"Failed to insert customer:\n{e}")
            return False
        
    def upload_file(self):
        filetypes = [("Excel files", "*.xlsx"), ("CSV files", "*.csv")]
        filename = filedialog.askopenfilename(title="Select Excel/CSV file", filetypes=filetypes)
        if not filename:
            return

        try:
            if filename.endswith(".xlsx"):
                wb = load_workbook(filename)
                ws = wb.active
                data = [row for row in ws.iter_rows(min_row=2, values_only=True)]
            else:
                with open(filename, newline='') as csvfile:
                    reader = csv.reader(csvfile)
                    next(reader)  # Skip header
                    data = [row for row in reader]

            count = 0
            for row in data:
                if len(row) < 7:
                    continue  # Skip incomplete rows

                cust_name, mobile, dob, pan, email, address, state_code = row[:7]

                # ✅ Get state name from Flask API
                state_name = self.get_state_name_by_code(state_code)
                if not state_name:
                    continue  # Skip if state code is invalid

                cust_id = self.generate_customer_id()  # Assume this is defined
                customer_payload = {
                    "cust_id": cust_id,
                    "cust_name": cust_name,
                    "mobile": mobile,
                    "dob": dob,
                    "pan": pan,
                    "email": email,
                    "address": address,
                    "state_code": state_code,
                    "state_name": state_name
                }

                # ✅ Insert using Flask API
                if self.insert_customer_api(customer_payload):
                    count += 1

            messagebox.showinfo("Upload Success", f"{count} customers uploaded successfully.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to upload file:\n{e}")

           
if __name__ == "__main__":
    root = tk.Tk()
    app = CustomerApp(root)
    root.mainloop()